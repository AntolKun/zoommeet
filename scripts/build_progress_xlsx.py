"""
Generates project_progress.xlsx — a snapshot of what's been built vs. not yet.

Sheets:
  - Overview          : summary counts per category
  - Sudah Dibuat      : full list of completed features (backend + frontend + infra)
  - Belum Dibuat      : full gap list, organized by section (matches Zoom/Meet comparison)
  - Roadmap           : sprint-mapped view
"""

import io
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# ---------- Styling -----------
HEADER_FILL = PatternFill("solid", fgColor="1c1917")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="f5efe9")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="847972")

THIN = Side(border_style="thin", color="3f3833")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CAT_FILL = {
    "Backend":  PatternFill("solid", fgColor="fef3c7"),
    "Frontend": PatternFill("solid", fgColor="dbeafe"),
    "Infra":    PatternFill("solid", fgColor="e0e7ff"),
    "Docs":     PatternFill("solid", fgColor="f3e8ff"),
    "Both":     PatternFill("solid", fgColor="ecfccb"),
}

DONE_FILL = PatternFill("solid", fgColor="dcfce7")
WIP_FILL  = PatternFill("solid", fgColor="fef9c3")
TBD_FILL  = PatternFill("solid", fgColor="fee2e2")


def style_header_row(ws, row: int, last_col: int):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths: dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------- DATA: Sudah Dibuat -----------
done_rows = [
    # (Kategori, Sprint/Tahap, Fitur, Bagian, File path utama, Catatan)
    # ===== Sprint 1 — Foundation =====
    ("Infra",    "Sprint 1", "Docker Compose: LiveKit + MySQL + Redis",                 "Infra",    "docker-compose.yml", "Awal project"),
    ("Backend",  "Sprint 1", "Go module + struktur internal/",                          "Backend",  "backend/go.mod", "Gin + middleware skeleton"),
    ("Backend",  "Sprint 1", "Config loader (.env via godotenv)",                       "Backend",  "backend/internal/config/config.go", ""),
    ("Backend",  "Sprint 1", "CORS middleware (dev permissive, prod env-driven)",       "Backend",  "backend/internal/middleware/cors.go", "AllowOriginFunc di dev"),
    ("Backend",  "Sprint 1", "GET /api/health endpoint",                                "Backend",  "backend/cmd/server/main.go", ""),
    ("Backend",  "Sprint 1", "POST /api/token (LiveKit JWT generation)",                "Backend",  "backend/internal/handlers/token.go", "Pakai livekit/protocol/auth"),

    # ===== Sprint 2 — Auth & DB =====
    ("Backend",  "Sprint 2", "Migration 001: users + rooms tables",                     "Backend",  "db/migrations/001_init.sql", ""),
    ("Backend",  "Sprint 2", "MySQL connection (database/sql + pool tuning)",           "Backend",  "backend/internal/db/db.go", "MaxOpenConns=25"),
    ("Backend",  "Sprint 2", "User model + UserRepo (Create, GetByID, GetByEmail)",     "Backend",  "backend/internal/repo/user_repo.go", ""),
    ("Backend",  "Sprint 2", "App JWT helper (golang-jwt/v5)",                          "Backend",  "backend/internal/auth/jwt.go", "HS256, 24h TTL"),
    ("Backend",  "Sprint 2", "RequireAuth middleware (Bearer token)",                   "Backend",  "backend/internal/middleware/auth.go", ""),
    ("Backend",  "Sprint 2", "POST /api/auth/register (bcrypt hash)",                   "Backend",  "backend/internal/handlers/auth.go", ""),
    ("Backend",  "Sprint 2", "POST /api/auth/login",                                    "Backend",  "backend/internal/handlers/auth.go", ""),
    ("Backend",  "Sprint 2", "Rate limit middleware (token bucket per IP)",             "Backend",  "backend/internal/middleware/ratelimit.go", "5 burst/min — di /auth endpoints"),

    # ===== Sprint 3 — Room Management =====
    ("Backend",  "Sprint 3", "Room model + RoomRepo (CRUD)",                            "Backend",  "backend/internal/repo/room_repo.go", ""),
    ("Backend",  "Sprint 3", "POST /api/rooms (create)",                                "Backend",  "backend/internal/handlers/rooms.go", ""),
    ("Backend",  "Sprint 3", "GET /api/rooms/my (list owned)",                          "Backend",  "backend/internal/handlers/rooms.go", ""),
    ("Backend",  "Sprint 3", "GET /api/rooms/:idOrSlug (detail)",                       "Backend",  "backend/internal/handlers/rooms.go", ""),
    ("Backend",  "Sprint 3", "DELETE /api/rooms/:idOrSlug",                             "Backend",  "backend/internal/handlers/rooms.go", "Owner only, cascading delete"),
    ("Backend",  "Sprint 3", "Token endpoint access control (public/private/owner)",   "Backend",  "backend/internal/handlers/token.go", ""),

    # ===== Sprint 4 — Meeting Features Lanjutan =====
    ("Backend",  "Sprint 4", "Migration 002: messages table (chat history)",            "Backend",  "db/migrations/002_messages.sql", ""),
    ("Backend",  "Sprint 4", "Migration 003: is_locked column",                          "Backend",  "db/migrations/003_room_lock.sql", ""),
    ("Backend",  "Sprint 4", "Migration 004: recordings table",                         "Backend",  "db/migrations/004_recordings.sql", ""),
    ("Backend",  "Sprint 4", "MessageRepo + chat endpoints (POST/GET messages)",        "Backend",  "backend/internal/handlers/chat.go", "Cursor pagination via 'before'"),
    ("Backend",  "Sprint 4", "Host controls: lock/unlock room",                          "Backend",  "backend/internal/handlers/participants.go", "API only"),
    ("Backend",  "Sprint 4", "Host controls: list/mute/kick participants",              "Backend",  "backend/internal/handlers/participants.go", "Pakai LiveKit RoomService"),
    ("Backend",  "Sprint 4", "LiveKit RoomService client wrapper",                      "Backend",  "backend/internal/livekit/client.go", ""),
    ("Backend",  "Sprint 4", "LiveKit Egress client wrapper",                           "Backend",  "backend/internal/livekit/egress.go", ""),
    ("Backend",  "Sprint 4", "Recording API (start/stop/list/get)",                     "Backend",  "backend/internal/handlers/recordings.go", "S3 storage config dependent"),
    ("Infra",    "Sprint 4", "MinIO + Egress container di docker-compose",              "Infra",    "docker-compose.yml + config/", ""),
    ("Infra",    "Sprint 4", "LiveKit config YAML (non-dev, Redis-enabled)",            "Infra",    "config/livekit.yaml", ""),

    # ===== Production Hygiene Quick Wins =====
    ("Backend",  "Polish",   "Gin release mode via APP_ENV=production",                 "Backend",  "backend/cmd/server/main.go", ""),
    ("Backend",  "Polish",   "Graceful shutdown (SIGINT/SIGTERM + ctx timeout)",        "Backend",  "backend/cmd/server/main.go", ""),

    # ===== Docs =====
    ("Docs",     "Polish",   "API reference Markdown",                                  "Docs",     "backend/API.md", "Lengkap dengan request/response + error codes"),
    ("Docs",     "Polish",   "Swagger UI di /swagger/index.html",                       "Docs",     "backend/docs/", "swaggo/gin-swagger; dev only"),
    ("Docs",     "Polish",   "Swag annotations di semua handler",                       "Docs",     "backend/internal/handlers/*.go", "swag init regen ulang"),

    # ===== Guest Mode =====
    ("Backend",  "Guest",    "POST /rooms/:slug/guest-token (no auth)",                 "Backend",  "backend/internal/handlers/token.go", "Random guest identity"),
    ("Backend",  "Guest",    "Room route public (di luar ProtectedRoute)",              "Frontend", "frontend/src/App.tsx", ""),

    # ===== Sprint Frontend =====
    ("Frontend", "FE Hari 1","Vite + React 19 + TypeScript scaffold",                  "Frontend", "frontend/vite.config.ts", ""),
    ("Frontend", "FE Hari 1","Tailwind v4 (config-less via plugin)",                    "Frontend", "frontend/src/index.css", "@import 'tailwindcss'"),
    ("Frontend", "FE Hari 1","React Router v7 + TanStack Query v5",                     "Frontend", "frontend/src/App.tsx + main.tsx", ""),
    ("Frontend", "FE Hari 1","Path alias @/* (vite + tsconfig)",                       "Frontend", "frontend/vite.config.ts", ""),
    ("Frontend", "FE Hari 1","Folder structure (pages/components/lib/hooks)",          "Frontend", "frontend/src/", ""),
    ("Frontend", "FE Hari 1","Env config (VITE_API_BASE_URL)",                          "Frontend", "frontend/.env", ""),
    ("Frontend", "FE Hari 1","API wrapper + JWT in localStorage",                       "Frontend", "frontend/src/lib/api.ts", "ApiError class + code field"),
    ("Frontend", "FE Hari 1","useAuth hook (token state)",                              "Frontend", "frontend/src/hooks/useAuth.ts", ""),
    ("Frontend", "FE Hari 1","Routing skeleton 7 routes + ProtectedRoute + NotFound",   "Frontend", "frontend/src/App.tsx", ""),

    ("Frontend", "FE Hari 2","Custom theme: zinc warm + orange flame",                  "Frontend", "frontend/src/index.css", "@theme di Tailwind v4"),
    ("Frontend", "FE Hari 2","Plus Jakarta Sans + JetBrains Mono fonts",                "Frontend", "frontend/src/index.css", "@fontsource-variable"),
    ("Frontend", "FE Hari 2","Form primitives: Button, Field, Input, PasswordInput, Alert","Frontend","frontend/src/components/ui/", ""),
    ("Frontend", "FE Hari 2","Layout (header + WIB clock + footer)",                    "Frontend", "frontend/src/components/Layout.tsx", ""),
    ("Frontend", "FE Hari 2","AuthLayout (split layout)",                               "Frontend", "frontend/src/components/AuthLayout.tsx", ""),
    ("Frontend", "FE Hari 2","Login page (react-hook-form + zod + API)",                "Frontend", "frontend/src/pages/Login.tsx", ""),
    ("Frontend", "FE Hari 2","Register page",                                           "Frontend", "frontend/src/pages/Register.tsx", ""),
    ("Frontend", "FE Hari 2","Home/landing page (asymmetric + mock meeting card)",      "Frontend", "frontend/src/pages/Home.tsx", ""),
    ("Frontend", "FE Hari 2","Film-grain noise overlay + warm focus ring",              "Frontend", "frontend/src/index.css", "UI/UX human, anti-AI generic"),

    ("Frontend", "FE Hari 3","useRooms hook (TanStack Query list/create/delete)",       "Frontend", "frontend/src/hooks/useRooms.ts", ""),
    ("Frontend", "FE Hari 3","Badge + Dialog primitive components",                     "Frontend", "frontend/src/components/ui/", ""),
    ("Frontend", "FE Hari 3","CreateRoomDialog (name + slug + public toggle)",          "Frontend", "frontend/src/components/CreateRoomDialog.tsx", ""),
    ("Frontend", "FE Hari 3","copyText clipboard helper (fallback legacy)",             "Frontend", "frontend/src/lib/clipboard.ts", ""),
    ("Frontend", "FE Hari 3","Dashboard: list/empty/skeleton/copy link/delete-confirm", "Frontend", "frontend/src/pages/Dashboard.tsx", ""),

    ("Frontend", "FE Hari 4","Install @livekit/components-react + livekit-client",      "Frontend", "frontend/package.json", "Auto code-split via React.lazy"),
    ("Frontend", "FE Hari 4","useRoomToken hook (POST /api/token)",                     "Frontend", "frontend/src/hooks/useRoomToken.ts", ""),
    ("Frontend", "FE Hari 4","Lobby page (input slug → navigate)",                      "Frontend", "frontend/src/pages/Lobby.tsx", "Auto-extract slug dari URL lengkap"),
    ("Frontend", "FE Hari 4","Room page dengan phase state machine",                    "Frontend", "frontend/src/pages/Room.tsx", "prejoin → connecting → joined → disconnected"),
    ("Frontend", "FE Hari 4","Pre-join screen (mic/cam toggle)",                        "Frontend", "frontend/src/pages/Room.tsx", ""),
    ("Frontend", "FE Hari 4","Error mapping (404/403/401) dengan copy Indonesia",       "Frontend", "frontend/src/pages/Room.tsx", ""),
    ("Frontend", "FE Hari 4","<LiveKitRoom> + <VideoConference> integration",           "Frontend", "frontend/src/pages/Room.tsx", "Code-split chunk ~600kb"),

    # ===== Guest mode FE =====
    ("Frontend", "Guest",    "Pre-join username field untuk guest",                     "Frontend", "frontend/src/pages/Room.tsx", "Persist ke localStorage"),
    ("Frontend", "Guest",    "useRoomToken support guest mode",                          "Frontend", "frontend/src/hooks/useRoomToken.ts", ""),

    # ===== Chat toast notif =====
    ("Frontend", "Polish",   "ChatToastNotifier (bottom-right fade in/out)",            "Frontend", "frontend/src/components/ChatToastNotifier.tsx", "Pakai useChat LiveKit"),

    # ===== Disconnect Recovery =====
    ("Frontend", "Polish",   "Disconnect → 'Sambung lagi' UI (gak auto-navigate)",      "Frontend", "frontend/src/pages/Room.tsx", "DisconnectedState component"),
    ("Frontend", "Polish",   "Auto-rejoin exponential backoff (3 retries: 1.5s/3s/6s)", "Frontend", "frontend/src/pages/Room.tsx", "Skip terminal reasons"),

    # ===== Recording UI =====
    ("Frontend", "Recording","useRoomInfo hook (room detail + isOwner)",                "Frontend", "frontend/src/hooks/useRoomInfo.ts", ""),
    ("Frontend", "Recording","getCurrentUserId helper (JWT decode payload)",            "Frontend", "frontend/src/lib/api.ts", ""),
    ("Frontend", "Recording","RecordingIndicator (badge buat semua peserta)",           "Frontend", "frontend/src/components/RecordingIndicator.tsx", "Pakai useIsRecording LiveKit"),
    ("Frontend", "Recording","RecordingControl (owner only, start/stop)",               "Frontend", "frontend/src/components/RecordingControl.tsx", "Auto-recover activeId via list endpoint"),

    # ===== Host Controls UI =====
    ("Frontend", "Host UI",  "RoomControls (floating toolbar top-left)",                "Frontend", "frontend/src/components/RoomControls.tsx", ""),
    ("Frontend", "Host UI",  "ParticipantsPanel (slide-in dari kiri + mute/kick)",      "Frontend", "frontend/src/components/ParticipantsPanel.tsx", "Owner only"),

    # ===== Reactions =====
    ("Frontend", "Engage",   "Reactions (8 emoji picker via data channel)",             "Frontend", "frontend/src/components/Reactions.tsx", "Topic 'vc.reaction'"),
    ("Frontend", "Engage",   "Floating emoji animation (CSS keyframe)",                 "Frontend", "frontend/src/index.css", ".reaction-float, 2.8s"),

    # ===== Raise Hand =====
    ("Frontend", "Engage",   "RaiseHandButton (toggle via participant.setAttributes)",  "Frontend", "frontend/src/components/RaiseHandButton.tsx", ""),
    ("Frontend", "Engage",   "ParticipantsPanel ✋ indicator + sort + counter badge",   "Frontend", "frontend/src/components/ParticipantsPanel.tsx", ""),

    # ===== Sound Effects =====
    ("Frontend", "Engage",   "Web Audio API sound helpers (join/leave/chat/reaction)",  "Frontend", "frontend/src/lib/sounds.ts", "No asset files, mute toggle localStorage"),
    ("Frontend", "Engage",   "RoomEventSounds (LiveKit event → play)",                  "Frontend", "frontend/src/components/RoomEventSounds.tsx", ""),

    # ===== Tunnel / Deploy =====
    ("Infra",    "Deploy",   "Vite preview production build (no HMR)",                  "Infra",    "frontend/vite.config.ts", "Production-stable buat testing live"),
    ("Infra",    "Deploy",   "cloudflared free tunnel buat URL HTTPS publik",           "Infra",    "(runtime)", "URL ganti tiap restart"),
    ("Infra",    "Deploy",   "LiveKit Cloud integration (free tier)",                   "Infra",    "backend/.env", "Ganti 3 env doang, code provider-agnostic"),

    # ===== Scheduling =====
    ("Backend",  "Sched",    "Migration 005: scheduled_at + duration_minutes",          "Backend",  "db/migrations/005_room_schedule.sql", ""),
    ("Backend",  "Sched",    "CreateRoom accept scheduled_at + duration",                "Backend",  "backend/internal/handlers/rooms.go", "Validation 5-480 min"),
    ("Frontend", "Sched",    "schedule.ts helpers (format + relative)",                 "Frontend", "frontend/src/lib/schedule.ts", "scheduleRelative dengan upcoming/live/past"),
    ("Frontend", "Sched",    "CreateRoomDialog toggle 'Dijadwalkan' + datetime + durasi","Frontend","frontend/src/components/CreateRoomDialog.tsx", ""),
    ("Frontend", "Sched",    "Dashboard schedule display + sort + auto-refresh 30s",    "Frontend", "frontend/src/pages/Dashboard.tsx", ""),
    ("Frontend", "Sched",    "Schedule badge (terjadwal/live/selesai)",                 "Frontend", "frontend/src/pages/Dashboard.tsx", ""),

    # ===== Password Room =====
    ("Backend",  "Password", "Migration 006: password_hash column",                     "Backend",  "db/migrations/006_room_password.sql", "bcrypt"),
    ("Backend",  "Password", "Token + GuestToken cek password (owner bypass)",          "Backend",  "backend/internal/handlers/token.go", "Error code password_required/invalid"),
    ("Frontend", "Password", "ApiError.code field",                                      "Frontend", "frontend/src/lib/api.ts", ""),
    ("Frontend", "Password", "CreateRoomDialog field password",                          "Frontend", "frontend/src/components/CreateRoomDialog.tsx", ""),
    ("Frontend", "Password", "Pre-join auto-prompt password (saat password_required)",  "Frontend", "frontend/src/pages/Room.tsx", ""),
    ("Frontend", "Password", "Dashboard badge 🔒 password",                              "Frontend", "frontend/src/pages/Dashboard.tsx", ""),

    # ===== Invite =====
    ("Frontend", "Invite",   "lib/invite.ts (.ics generator RFC 5545 + mailto)",        "Frontend", "frontend/src/lib/invite.ts", "Pure client, no SMTP"),
    ("Frontend", "Invite",   "Dashboard Share menu (Copy/Calendar/Email)",              "Frontend", "frontend/src/pages/Dashboard.tsx", ""),

    # ===== Recurring =====
    ("Backend",  "Recurring","Migration 007: recurrence column",                        "Backend",  "db/migrations/007_room_recurrence.sql", "daily/weekly"),
    ("Backend",  "Recurring","CreateRoom validate recurrence (requires scheduled_at)",  "Backend",  "backend/internal/handlers/rooms.go", ""),
    ("Frontend", "Recurring","nextOccurrence + recurrenceLabel helpers",                "Frontend", "frontend/src/lib/schedule.ts", ""),
    ("Frontend", "Recurring","Dialog dropdown 'Pengulangan'",                            "Frontend", "frontend/src/components/CreateRoomDialog.tsx", ""),
    ("Frontend", "Recurring","Dashboard 🔁 + 'Setiap Senin · 14.00'",                   "Frontend", "frontend/src/pages/Dashboard.tsx", ""),
    ("Frontend", "Recurring","Sort by next occurrence (always 'upcoming')",             "Frontend", "frontend/src/pages/Dashboard.tsx", ""),

    # ===== Waiting Room (in progress) =====
    ("Backend",  "Waiting",  "Migration 008: waiting_room_enabled + waiting_requests",  "Backend",  "db/migrations/008_waiting_room.sql", "⚠️ IN PROGRESS"),
    ("Backend",  "Waiting",  "WaitingRequest model",                                     "Backend",  "backend/internal/models/waiting.go", "⚠️ IN PROGRESS"),
    ("Backend",  "Waiting",  "WaitingRepo (Create/GetByToken/ListPending/Approve/Deny)", "Backend",  "backend/internal/repo/waiting_repo.go", "⚠️ IN PROGRESS"),
    ("Backend",  "Waiting",  "Room.WaitingRoomEnabled field + repo update",              "Backend",  "backend/internal/models/room.go + repo", "⚠️ IN PROGRESS"),
]

# ---------- DATA: Belum Dibuat -----------
notyet_rows = [
    # (Section, Fitur, Effort S/M/L, Catatan)
    # Section 1: Sebelum meeting
    ("01. Sebelum meeting", "Waiting room (sisanya: handlers + token integration + frontend UI)", "M", "Sebagian backend udah; sisa: tokenResponse pakai status field, handler status/list/admit/deny/toggle, routes, frontend hooks + Room phase + WaitingRoomPanel + dialog toggle"),
    ("01. Sebelum meeting", "Co-host designation", "M", "Tabel room_cohosts + 3 endpoint + permission helper diterapkan ke ~10 endpoint owner-gated (lock/mute/kick/recording)"),
    ("01. Sebelum meeting", "Send invitation email via SMTP", "M", "Sekarang mailto: client-only. Butuh SMTP service (Brevo/SES/Gmail App Pass) + worker"),
    ("01. Sebelum meeting", "Google Calendar / Outlook OAuth", "L", "Full OAuth flow + token storage + event create API"),

    # Section 2: Pre-join
    ("02. Pre-join screen", "Live camera preview", "S", "getUserMedia + <video> di pre-join"),
    ("02. Pre-join screen", "Mic level meter (visualisasi audio)", "S", "Web Audio AnalyserNode di stream"),
    ("02. Pre-join screen", "Device picker (camera/mic/speaker)", "S", "navigator.mediaDevices.enumerateDevices + select"),
    ("02. Pre-join screen", "Test audio (record-playback)", "S", "Sederhana: rekam ~3s, playback"),
    ("02. Pre-join screen", "Saved device preferences", "S", "localStorage"),
    ("02. Pre-join screen", "Background blur preview", "M", "LiveKit Track Processor"),
    ("02. Pre-join screen", "Virtual background preview", "L", "TensorFlow.js/MediaPipe segmentation"),

    # Section 3: Layout
    ("03. Tampilan & layout", "Speaker view vs grid toggle", "S", "Ganti VideoConference layout prop"),
    ("03. Tampilan & layout", "Pin participant (untuk self)", "S", ""),
    ("03. Tampilan & layout", "Spotlight (host pin ke semua)", "M", "Pakai room metadata"),
    ("03. Tampilan & layout", "Hide self view", "S", ""),
    ("03. Tampilan & layout", "Picture-in-Picture browser", "S", "document.pictureInPictureElement"),
    ("03. Tampilan & layout", "Full-screen toggle button", "S", "requestFullscreen()"),
    ("03. Tampilan & layout", "Timer durasi meeting di corner", "S", ""),
    ("03. Tampilan & layout", "Connection quality indicator per participant", "S", "LiveKit Participant.connectionQuality"),
    ("03. Tampilan & layout", "Floating self-view drag", "M", ""),

    # Section 4: Audio
    ("04. Audio", "Push-to-talk (tahan spasi)", "S", ""),
    ("04. Audio", "'Kamu di-mute tapi sedang bicara' warning", "S", "VAD via Web Audio"),
    ("04. Audio", "Audio-only mode (matiin video)", "S", ""),
    ("04. Audio", "Noise suppression toggle", "M", "LiveKit Cloud support"),
    ("04. Audio", "Music mode", "S", ""),
    ("04. Audio", "Live transcription / captions", "L", "Whisper API atau LiveKit AI agent"),

    # Section 5: Video
    ("05. Video", "Mirror video / flip", "S", ""),
    ("05. Video", "HD video toggle", "S", "LiveKit publishOptions"),
    ("05. Video", "Avatar di tile kalo cam off (selain inisial)", "S", ""),
    ("05. Video", "Profile picture upload + display", "M", "Storage S3 + UI dialog"),
    ("05. Video", "Background blur", "M", ""),
    ("05. Video", "Virtual background", "L", ""),
    ("05. Video", "Beauty filter / AR effects", "L", ""),

    # Section 6: Screen share
    ("06. Screen share", "Share with audio (Chrome tab)", "S", "getDisplayMedia({audio:true})"),
    ("06. Screen share", "Stop participant's share (host)", "S", "Pakai mute API"),
    ("06. Screen share", "Allow/disallow share (room setting)", "S", "Tambah field permission di token"),
    ("06. Screen share", "Laser pointer", "S", "Cursor overlay synced via data channel"),
    ("06. Screen share", "Annotation tools (gambar di screen)", "M", "Canvas overlay + sync"),
    ("06. Screen share", "Whiteboard collaborative", "L", "Canvas yang shared via data channel"),

    # Section 7: Chat
    ("07. Chat", "Pakai chat history persistent kita (sync ke backend)", "M", "Sekarang LiveKit data channel doang, gak ke DB"),
    ("07. Chat", "Private 1-on-1 chat", "M", "Tambah recipient_id, ubah UI"),
    ("07. Chat", "@mention dengan notification", "S", ""),
    ("07. Chat", "Emoji picker di input", "S", ""),
    ("07. Chat", "Edit / delete pesan sendiri", "S", "Backend endpoint + UI"),
    ("07. Chat", "Reactions ke pesan (like, dll)", "S", ""),
    ("07. Chat", "Reply / thread", "M", ""),
    ("07. Chat", "File / image attachment", "M", "Storage S3 + UI uploader"),

    # Section 8: Recording (UI done, infra pending)
    ("08. Recording", "Banner consent 'Meeting direkam' menonjol", "S", ""),
    ("08. Recording", "Download link recording setelah selesai", "S", "Tergantung S3"),
    ("08. Recording", "Storage S3 setup di LiveKit Cloud", "M", "Konfigurasi external (bukan code)"),
    ("08. Recording", "Pilih layout recording (grid/speaker/share)", "S", "Param Egress request"),
    ("08. Recording", "Pause/resume", "M", "LiveKit Egress API"),
    ("08. Recording", "Auto-transcribe", "L", ""),

    # Section 9: Host controls (API exists, mostly UI work)
    ("09. Host controls", "Lock room toggle di toolbar (UI)", "S", "API udah ada"),
    ("09. Host controls", "Mute all participants sekaligus", "S", "Loop ke semua + call mute API"),
    ("09. Host controls", "Lower all hands sekaligus", "S", ""),
    ("09. Host controls", "Allow/disallow participants unmute themselves", "M", "Permission flag di room"),
    ("09. Host controls", "Disable chat (room setting)", "S", ""),
    ("09. Host controls", "Disable rename", "S", ""),

    # Section 10: Participant interactions
    ("10. Participant", "Yes/No quick reaction (beda dari emoji)", "S", ""),
    ("10. Participant", "Profile picture/avatar untuk akun login", "M", "S3 upload + UI"),
    ("10. Participant", "Display name change on-the-fly", "S", "setName via LiveKit"),

    # Section 11: Breakout rooms
    ("11. Breakout rooms", "Bikin sub-rooms breakout", "L", "Sub-room creation + assign"),
    ("11. Breakout rooms", "Auto / manual assign peserta", "M", ""),
    ("11. Breakout rooms", "Broadcast pesan ke semua breakout", "M", ""),
    ("11. Breakout rooms", "Recall everyone ke main room", "M", ""),

    # Section 12: Polls / Q&A / Tools
    ("12. Polls / Q&A", "Live polling (host kasih, peserta vote)", "M", "Tabel polls + votes + UI"),
    ("12. Polls / Q&A", "Q&A panel terstruktur", "M", ""),
    ("12. Polls / Q&A", "Quiz mode", "L", ""),
    ("12. Polls / Q&A", "Whiteboard interaktif", "L", ""),

    # Section 13: Connection / network
    ("13. Connection", "Bandwidth/quality indicator (RTT/jitter/loss)", "S", "LiveKit Room.engine stats"),
    ("13. Connection", "Manual quality switch (Auto/Low/Med/High)", "S", "publishOptions per quality"),
    ("13. Connection", "Reconnect status banner enhanced", "S", "Lebih informatif"),
    ("13. Connection", "Show statistics debug panel", "S", ""),
    ("13. Connection", "Network change detection (pindah WiFi)", "S", ""),

    # Section 14: Mobile
    ("14. Mobile", "Capacitor wrap (APK/IPA)", "M", "Per rencana awal — perlu DNS + cert"),
    ("14. Mobile", "Push notification incoming meeting", "L", "FCM/APNs + SDK"),
    ("14. Mobile", "Background audio mode (lock screen)", "S", ""),
    ("14. Mobile", "Wake lock (layar tidak mati)", "S", "Wake Lock API"),
    ("14. Mobile", "PSTN dial-in (telpon biasa)", "L", "LiveKit SIP / vendor (Twilio)"),

    # Section 15: Accessibility
    ("15. Accessibility", "Closed captions / live captions", "L", "Whisper/STT integration"),
    ("15. Accessibility", "Live transcription panel", "L", ""),
    ("15. Accessibility", "Keyboard shortcuts (mute/cam/leave)", "S", ""),
    ("15. Accessibility", "Shortcut cheat sheet (tekan '?')", "S", ""),
    ("15. Accessibility", "High contrast theme", "S", ""),
    ("15. Accessibility", "Light mode toggle", "S", "Sekarang dark only"),

    # Section 16: Notifications / sounds
    ("16. Notifications", "Sound mute toggle button (UI)", "S", "Logic ada di lib/sounds.ts"),
    ("16. Notifications", "Browser notification kalo tab background", "S", "Notification API"),
    ("16. Notifications", "'Yakin keluar?' confirmation sebelum disconnect", "S", ""),
    ("16. Notifications", "Granular toggle per sound type (di settings)", "S", ""),

    # Section 17: Security
    ("17. Security", "End-to-end encryption (E2EE)", "L", "LiveKit support — config + key exchange"),
    ("17. Security", "Watermark di shared screen", "M", "Canvas overlay"),
    ("17. Security", "Disable copy chat", "S", "user-select:none + JS"),
    ("17. Security", "Required login domain (mis. @piko.co.id)", "S", "Cek di register/login"),

    # Section 18: Admin / analytics
    ("18. Admin", "Meeting attendance report", "M", "Log participants per session"),
    ("18. Admin", "Meeting duration / quality stats", "M", ""),
    ("18. Admin", "Admin dashboard org-level", "L", ""),
    ("18. Admin", "Audit log host actions", "S", ""),
    ("18. Admin", "Usage chart sederhana", "S", ""),

    # Section 19: UX touches
    ("19. UX touches", "Settings dialog modal (sentral preferences)", "M", "Pintu masuk ke banyak settings"),
    ("19. UX touches", "Tooltip di tiap tombol", "S", ""),
    ("19. UX touches", "First-time tour / onboarding", "M", "Joyride atau custom"),
    ("19. UX touches", "Permission denied helper (kasih cara fix)", "S", ""),
    ("19. UX touches", "Default mute on join setting per user", "S", "localStorage"),
    ("19. UX touches", "Profile photo upload buat akun", "M", "S3"),
    ("19. UX touches", "Language selector (id/en)", "M", "i18next setup"),
    ("19. UX touches", "Show waktu lokal masing-masing peserta", "S", ""),

    # Sprint 5 — Fitur pembeda (belum decide)
    ("S5. Fitur pembeda", "Belum decide fitur diferensiator", "—", "Per xlsx: pilih 1-2 dari Waiting Room / AI Transcription / Breakout / Polling / Meeting Summary / Whiteboard"),

    # Sprint 6 — Production deploy
    ("S6. Deploy", "Deploy ke server beneran (VPS atau piko shared)", "L", "Sprint 6 territory"),
    ("S6. Deploy", "Production domain + HTTPS (Caddy/Let's Encrypt)", "M", ""),
    ("S6. Deploy", "Self-host LiveKit (atau LiveKit Cloud paid tier)", "M", "Tinggal flip 3 env"),
    ("S6. Deploy", "Monitoring (Sentry + Uptime Kuma)", "M", ""),
    ("S6. Deploy", "TURN server config (buat NAT traversal)", "M", "LiveKit built-in TURN"),
    ("S6. Deploy", "Database hardening + backup automated", "M", ""),
    ("S6. Deploy", "Production CORS allowlist (bukan permissive)", "S", "Tinggal env CORS_ORIGINS"),
    ("S6. Deploy", "Named cloudflared tunnel (URL stabil)", "S", "Butuh Cloudflare account"),
]


# ---------- DATA: Roadmap (sprint mapping) -----------
roadmap_rows = [
    ("Sprint 1 — Foundation",     "✅ Selesai", "Infra + backend health + LiveKit token + Frontend Hari 1-4 (auth, dashboard, video call basic)"),
    ("Sprint 2 — Auth & DB",      "✅ Selesai", "Migrations, user CRUD, JWT, register/login, rate limit"),
    ("Sprint 3 — Room Mgmt",      "✅ Selesai", "Rooms CRUD, access control, slug generation, dashboard UI"),
    ("Sprint 4 — Meeting Lanjut", "✅ Sebagian besar selesai", "Chat persist, host controls (API+UI), recording (API+UI, storage S3 belum disetup). UI: lock toggle belum"),
    ("Sprint 5 — Fitur Pembeda",  "⏳ Belum mulai", "Sebagian fitur sebenarnya udah ada (reactions, raise hand) tapi belum decide fitur HERO diferensiator"),
    ("Sprint 6 — Production Deploy","⏳ Belum mulai", "Server live, domain HTTPS, monitoring, TURN, hardening. Nunggu server Docker baru per info user"),
    ("Capacitor mobile wrap",     "⏳ Belum mulai", "Setelah Sprint 6 atau pas web udah polished"),
]


# ---------- BUILD WORKBOOK -----------
wb = Workbook()

# ----- Sheet 1: Overview -----
ws = wb.active
ws.title = "Overview"

ws["A1"] = "Videoconf App — Progress Snapshot"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Generated by scripts/build_progress_xlsx.py"
ws["A2"].font = SUBTITLE_FONT

# Counts by category
done_by_cat = {}
for cat, *_ in done_rows:
    done_by_cat[cat] = done_by_cat.get(cat, 0) + 1

ws["A4"] = "Total fitur sudah dibuat"
ws["A4"].font = Font(bold=True)
ws["B4"] = len(done_rows)

ws["A5"] = "Total gap (belum dibuat)"
ws["A5"].font = Font(bold=True)
ws["B5"] = len(notyet_rows)

ws["A6"] = "Persentase progress (kasar, by jumlah item)"
ws["A6"].font = Font(bold=True)
total = len(done_rows) + len(notyet_rows)
ws["B6"] = f"{round(100 * len(done_rows) / total)}%"

# Per-category breakdown
ws["A8"] = "Breakdown sudah dibuat per kategori"
ws["A8"].font = Font(bold=True)

ws["A9"] = "Kategori"
ws["B9"] = "Jumlah"
ws["C9"] = "Persen total done"
style_header_row(ws, 9, 3)

row = 10
for cat, count in sorted(done_by_cat.items(), key=lambda kv: -kv[1]):
    ws.cell(row=row, column=1, value=cat).fill = CAT_FILL.get(cat, PatternFill())
    ws.cell(row=row, column=2, value=count)
    ws.cell(row=row, column=3, value=f"{round(100 * count / len(done_rows))}%")
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = BORDER
    row += 1

# Counts by effort for not-yet
ws[f"A{row + 2}"] = "Sisa gap by effort estimate"
ws[f"A{row + 2}"].font = Font(bold=True)

effort_counts = {"S": 0, "M": 0, "L": 0, "—": 0}
for _, _, e, _ in notyet_rows:
    effort_counts[e] = effort_counts.get(e, 0) + 1

ws[f"A{row + 3}"] = "Effort"
ws[f"B{row + 3}"] = "Jumlah"
ws[f"C{row + 3}"] = "Catatan"
style_header_row(ws, row + 3, 3)

efforts_desc = {
    "S": "Small (<2 jam)",
    "M": "Medium (~setengah hari sampai 1 hari)",
    "L": "Large (>1 hari, multi-session)",
    "—": "Belum di-scope / butuh keputusan dulu",
}
r = row + 4
for k in ["S", "M", "L", "—"]:
    ws.cell(row=r, column=1, value=k)
    ws.cell(row=r, column=2, value=effort_counts.get(k, 0))
    ws.cell(row=r, column=3, value=efforts_desc[k])
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = BORDER
    r += 1

autosize(ws, {1: 40, 2: 14, 3: 50})

# ----- Sheet 2: Sudah Dibuat -----
ws = wb.create_sheet("Sudah Dibuat")
headers = ["Kategori", "Sprint / Tahap", "Fitur", "Bagian", "File path utama", "Catatan"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header_row(ws, 1, len(headers))

for r, row in enumerate(done_rows, start=2):
    for i, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=i, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if i == 4:
            cell.fill = CAT_FILL.get(val, PatternFill())
    # Status column visualization — light green fill on Fitur cell
    if "IN PROGRESS" in row[5]:
        ws.cell(row=r, column=3).fill = WIP_FILL
    else:
        ws.cell(row=r, column=3).fill = DONE_FILL

autosize(ws, {1: 12, 2: 12, 3: 56, 4: 10, 5: 50, 6: 60})
ws.freeze_panes = "A2"

# ----- Sheet 3: Belum Dibuat -----
ws = wb.create_sheet("Belum Dibuat")
headers = ["Section", "Fitur", "Effort", "Catatan"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header_row(ws, 1, len(headers))

for r, row in enumerate(notyet_rows, start=2):
    for i, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=i, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=r, column=2).fill = TBD_FILL

autosize(ws, {1: 26, 2: 56, 3: 10, 4: 70})
ws.freeze_panes = "A2"

# ----- Sheet 4: Roadmap -----
ws = wb.create_sheet("Roadmap")
headers = ["Sprint / Tahap", "Status", "Ringkasan"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header_row(ws, 1, len(headers))

for r, row in enumerate(roadmap_rows, start=2):
    for i, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=i, value=val)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if "✅" in row[1]:
        ws.cell(row=r, column=2).fill = DONE_FILL
    elif "⏳" in row[1]:
        ws.cell(row=r, column=2).fill = WIP_FILL

autosize(ws, {1: 32, 2: 24, 3: 100})
ws.freeze_panes = "A2"

# ----- Save -----
out = "project_progress.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"  Sudah dibuat : {len(done_rows)}")
print(f"  Belum dibuat : {len(notyet_rows)}")
print(f"  Progress kasar: {round(100 * len(done_rows) / (len(done_rows) + len(notyet_rows)))}%")
